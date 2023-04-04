import pandas as pd
import os
import json
import datetime
import xlsxwriter
import openpyxl
from datetime import date
import time
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from pytrends.request import TrendReq
pytrends = TrendReq()

def airlines():
    # Read in json object
    with open("data_queries.json", 'r') as f:
        data_queries_raw = json.load(f)

    # Create new object with only outdated queries (Last updated over 7 days ago)
    data_queries = {}
    for query in data_queries_raw.keys():
        if (datetime.date.today() - datetime.datetime.strptime(data_queries_raw[query]["last_updated"], "%Y-%m-%d").date()).days > 7:
            data_queries[query] = data_queries_raw[query]

    path = "RDI raw data.xlsx"
    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False

    def update_trend_data(query):
        xfile = openpyxl.load_workbook('RDI raw data.xlsx')
        sheet = xfile[query]
        # index of [sheet_name] sheet
        idx = xfile.sheetnames.index(query)
        # remove [sheet_name]
        xfile.remove(sheet)
        # create an empty sheet [sheet_name] using old index
        xfile.create_sheet(query, idx)
        sheet = xfile[query]
        sheet.append([])
        sheet.append([])
        pytrends.build_payload(kw_list=data_queries[query]["kw_list"],cat=0, timeframe="today 5-y", geo=data_queries[query]["geo"])
        df_raw = pytrends.interest_over_time()
        df = df_raw.reset_index()
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append([''] + r)
        xfile.save('RDI raw data.xlsx')
        # Update log file
        data_queries_raw[query]["last_updated"] = str(datetime.date.today().strftime("%Y-%m-%d"))
        with open('data_queries.json', 'w') as f:
            json.dump(data_queries_raw, f)
        print("{} was succesfully updated".format(query))
    
    # Get pytrends, update raw data file, save
    for query in data_queries.keys():
        try:
            update_trend_data(query)
        except:
            print("Last attempted query was {query} at {time}".format(query=query, time=datetime.datetime.now()))
            # time.sleep(500)
            


if __name__ == "__main__":
    pd.set_option("display.max_rows", None, "display.max_columns", None)
    airlines()
